"""Local Trials Queue dashboard.

This app is intentionally a local desktop UI rather than a hosted web server. It
reads and writes the shared Excel workbook directly, so teams can keep the
SharePoint-synced workbook as the source of truth.
"""

import json
import os
import threading
from copy import deepcopy
from datetime import datetime, timedelta, timezone
from pathlib import Path
from tkinter import BooleanVar, StringVar, Tk, messagebox, ttk
from tkinter.scrolledtext import ScrolledText
from uuid import uuid4

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL_PATH = BASE_DIR / "testing QUEUE.xlsx"
EXCEL_PATH = Path(os.environ.get("QUEUE_EXCEL_PATH", DEFAULT_EXCEL_PATH)).expanduser()
STATE_PATH = Path(os.environ.get("QUEUE_STATE_PATH", BASE_DIR / "data" / "queue_state.json")).expanduser()
CLAIM_TIMEOUT_SECONDS = int(os.environ.get("CLAIM_TIMEOUT_SECONDS", "120"))
POLL_SECONDS = int(os.environ.get("QUEUE_POLL_SECONDS", "10"))

READY_VALUES = {"x", "yes", "y", "true", "1"}
COMPLETE_VALUES = {"x", "yes", "y", "true", "1", "complete", "done"}
PRIORITY_ORDER = ["Expedites", "Funded Rentals", "Accessories", "Requested Ship Dates", "Regular Daily Queue"]

FIELD_ALIASES = {
    "id": ["ID"],
    "last_name": ["LastName", "Last Name"],
    "first_name": ["FirstName", "First Name"],
    "device": ["Device"],
    "loan_type": ["LoanType", "Loan Type"],
    "queue_date": ["QueueDate", "Queue Date"],
    "vocabulary": ["Vocabulary"],
    "notes": ["Notes"],
    "priority": ["Priority"],
    "ready_for_prep": ["ReadyForPrep", "Ready for Prep", "Ready", "Prepped?"],
    "prep_claimed_by": ["PrepClaimedBy", "Prep Claimed By"],
    "prep_claimed_at": ["PrepClaimedAt", "Prep Claimed At"],
    "prep_complete": ["PrepComplete", "Prep Complete"],
    "ready_for_qa": ["ReadyForQA", "Ready for QA", "ReadyFor2ndCheck", "Ready for 2nd Check"],
    "qa_claimed_by": ["QAClaimedBy", "QA Claimed By", "SecondCheckClaimedBy"],
    "qa_claimed_at": ["QAClaimedAt", "QA Claimed At", "SecondCheckClaimedAt"],
    "qa_complete": ["QAComplete", "QA Complete", "SecondCheckComplete"],
    "assignment_user": ["AssignmentUser", "Assignment User"],
    "assignment_expires": ["AssignmentExpires", "Assignment Expires"],
    "assignment_status": ["AssignmentStatus", "Assignment Status"],
}

EXCEL_LOCK = threading.Lock()
STATE_LOCK = threading.Lock()


def now_iso():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def truthy(value, complete=False):
    if value is None:
        return False
    values = COMPLETE_VALUES if complete else READY_VALUES
    return str(value).strip().lower() in values


def clean_user(user):
    user = (user or "").strip()
    if not user:
        raise ValueError("Enter your coordinator name or initials first.")
    return user


def normalize_queue(queue_name):
    queue_name = (queue_name or "prep").strip().lower()
    if queue_name in {"prep", "qa"}:
        return queue_name
    raise ValueError("Queue must be prep or qa.")


def priority_lane(row):
    text = " ".join(str(row.get(key) or "") for key in ("priority", "loan_type", "notes", "device", "queue_date")).lower()
    if "expedite" in text or "urgent" in text:
        return "Expedites"
    if "funded" in text or "rental" in text:
        return "Funded Rentals"
    if "accessor" in text:
        return "Accessories"
    if "requested ship" in text or "ship date" in text or "rsd" in text:
        return "Requested Ship Dates"
    return "Regular Daily Queue"


def sort_key(row):
    return (PRIORITY_ORDER.index(row["lane"]), str(row.get("sort_queue_date") or row.get("queue_date") or ""), row.get("row_number", 0))


def assignment_active(row):
    expires = row.get("assignment_expires")
    if not expires:
        return False
    try:
        if isinstance(expires, datetime):
            expires_at = expires if expires.tzinfo else expires.replace(tzinfo=timezone.utc)
        else:
            expires_at = datetime.fromisoformat(str(expires).replace("Z", "+00:00"))
        return expires_at > datetime.now(timezone.utc)
    except ValueError:
        return False


class ExcelQueueStore:
    def __init__(self, excel_path):
        self.excel_path = excel_path

    def _open(self):
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel queue file not found: {self.excel_path}")
        workbook = load_workbook(self.excel_path)
        worksheet = workbook.active
        headers = {str(cell.value).strip(): cell.column for cell in worksheet[1] if cell.value}
        field_map = {}
        for field, aliases in FIELD_ALIASES.items():
            for alias in aliases:
                if alias in headers:
                    field_map[field] = headers[alias]
                    break
        return workbook, worksheet, field_map

    def read_rows(self):
        with EXCEL_LOCK:
            workbook, worksheet, field_map = self._open()
            rows = []
            for row_number in range(2, worksheet.max_row + 1):
                if all(worksheet.cell(row_number, col).value in (None, "") for col in range(1, worksheet.max_column + 1)):
                    continue
                row = {field: worksheet.cell(row_number, col).value for field, col in field_map.items()}
                column_h_value = worksheet.cell(row_number, 8).value
                if not truthy(row.get("ready_for_prep")) and truthy(column_h_value):
                    row["ready_for_prep"] = column_h_value
                row["row_number"] = row_number
                row["display_name"] = f"{row.get('first_name') or ''} {row.get('last_name') or ''}".strip()
                row["lane"] = priority_lane(row)
                row["sort_queue_date"] = row.get("queue_date") or ""
                row["queue_date"] = self._display_value(row.get("queue_date"))
                row["prep_ready"] = truthy(row.get("ready_for_prep")) and not truthy(row.get("prep_complete"), complete=True)
                row["qa_ready"] = truthy(row.get("ready_for_qa")) and not truthy(row.get("qa_complete"), complete=True)
                row["assignment_active"] = assignment_active(row)
                rows.append(row)
            workbook.close()
            return rows

    def dashboard_rows(self):
        return sorted([row for row in self.read_rows() if row["prep_ready"] or row["qa_ready"]], key=sort_key)

    def add_row(self, payload):
        with EXCEL_LOCK:
            workbook, worksheet, field_map = self._open()
            next_row = worksheet.max_row + 1
            row_id = payload.get("id") or str(uuid4())[:8]
            values = {
                "id": row_id,
                "last_name": payload.get("last_name"),
                "first_name": payload.get("first_name"),
                "device": payload.get("device"),
                "loan_type": payload.get("loan_type"),
                "queue_date": payload.get("queue_date") or datetime.now().date().isoformat(),
                "vocabulary": payload.get("vocabulary"),
                "notes": payload.get("notes"),
                "priority": payload.get("priority"),
                "ready_for_prep": "X" if payload.get("ready_for_prep") else "",
                "prep_complete": "No",
                "ready_for_qa": "No",
                "qa_complete": "No",
                "assignment_status": "Ready" if payload.get("ready_for_prep") else "Draft",
            }
            for field, value in values.items():
                self._set_cell(worksheet, field_map, next_row, field, value)
            workbook.save(self.excel_path)
            workbook.close()
            return row_id

    def claim_order(self, row_number, queue_name, user):
        row = self._find_dashboard_row(row_number)
        if not row:
            raise ValueError("Order is no longer available.")
        expected_user = row.get("assignment_user")
        if row.get("assignment_active") and expected_user and expected_user != user:
            raise ValueError(f"This order is currently offered to {expected_user}.")
        stamp = now_iso()
        with EXCEL_LOCK:
            workbook, worksheet, field_map = self._open()
            if queue_name == "prep":
                self._set_cell(worksheet, field_map, row_number, "prep_claimed_by", user)
                self._set_cell(worksheet, field_map, row_number, "prep_claimed_at", stamp)
            else:
                self._set_cell(worksheet, field_map, row_number, "qa_claimed_by", user)
                self._set_cell(worksheet, field_map, row_number, "qa_claimed_at", stamp)
            self._set_cell(worksheet, field_map, row_number, "assignment_user", user)
            self._set_cell(worksheet, field_map, row_number, "assignment_status", "Claimed")
            self._set_cell(worksheet, field_map, row_number, "assignment_expires", "")
            workbook.save(self.excel_path)
            workbook.close()

    def complete_order(self, row_number, queue_name, user):
        stamp = now_iso()
        with EXCEL_LOCK:
            workbook, worksheet, field_map = self._open()
            if queue_name == "prep":
                self._set_cell(worksheet, field_map, row_number, "prep_claimed_by", user)
                self._set_cell(worksheet, field_map, row_number, "prep_claimed_at", stamp)
                self._set_cell(worksheet, field_map, row_number, "prep_complete", "Yes")
                self._set_cell(worksheet, field_map, row_number, "ready_for_qa", "Yes")
                status = "Ready for QA"
            else:
                self._set_cell(worksheet, field_map, row_number, "qa_claimed_by", user)
                self._set_cell(worksheet, field_map, row_number, "qa_claimed_at", stamp)
                self._set_cell(worksheet, field_map, row_number, "qa_complete", "Yes")
                status = "Complete"
            self._set_cell(worksheet, field_map, row_number, "assignment_user", user)
            self._set_cell(worksheet, field_map, row_number, "assignment_status", status)
            self._set_cell(worksheet, field_map, row_number, "assignment_expires", "")
            workbook.save(self.excel_path)
            workbook.close()

    def set_assignment(self, row_number, user, expires_at, status):
        with EXCEL_LOCK:
            workbook, worksheet, field_map = self._open()
            self._set_cell(worksheet, field_map, row_number, "assignment_user", user)
            self._set_cell(worksheet, field_map, row_number, "assignment_expires", expires_at)
            self._set_cell(worksheet, field_map, row_number, "assignment_status", status)
            workbook.save(self.excel_path)
            workbook.close()

    def clear_assignment(self, row_number, user, status):
        with EXCEL_LOCK:
            workbook, worksheet, field_map = self._open()
            self._set_cell(worksheet, field_map, row_number, "assignment_user", user)
            self._set_cell(worksheet, field_map, row_number, "assignment_expires", "")
            self._set_cell(worksheet, field_map, row_number, "assignment_status", status)
            workbook.save(self.excel_path)
            workbook.close()

    def _find_dashboard_row(self, row_number):
        return next((row for row in self.dashboard_rows() if row["row_number"] == row_number), None)

    @staticmethod
    def _display_value(value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        return str(value)

    @staticmethod
    def _set_cell(worksheet, field_map, row_number, field, value):
        if field in field_map:
            worksheet.cell(row_number, field_map[field]).value = value


class QueueState:
    def __init__(self, path):
        self.path = path

    def load(self):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if not self.path.exists():
            return {"prep": [], "qa": []}
        try:
            with self.path.open() as state_file:
                state = json.load(state_file)
        except json.JSONDecodeError:
            state = {"prep": [], "qa": []}
        for queue_name in ("prep", "qa"):
            state.setdefault(queue_name, [])
        return state

    def save(self, state):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self.path.open("w") as state_file:
            json.dump(state, state_file, indent=2)
        return deepcopy(state)

    def join(self, queue_name, user):
        queue_name = normalize_queue(queue_name)
        with STATE_LOCK:
            state = self.load()
            state[queue_name] = [entry for entry in state[queue_name] if entry["user"] != user]
            state[queue_name].append({"user": user, "joined_at": now_iso()})
            return self.save(state)

    def rotate_to_bottom(self, queue_name, user):
        queue_name = normalize_queue(queue_name)
        with STATE_LOCK:
            state = self.load()
            remaining = [entry for entry in state[queue_name] if entry["user"] != user]
            remaining.append({"user": user, "joined_at": now_iso()})
            state[queue_name] = remaining
            return self.save(state)

    def first_user(self, queue_name):
        state = self.load()
        queue = state[normalize_queue(queue_name)]
        return queue[0]["user"] if queue else None


class TrialsDashboard(Tk):
    def __init__(self, store, queues):
        super().__init__()
        self.store = store
        self.queues = queues
        self.title("Trials Queue Dashboard - Local Excel UI")
        self.geometry("1240x760")
        self.selected_rows = {}
        self.notified_offers = set()
        self.user_name = StringVar()
        self._build_ui()
        self.refresh_all()
        self.after(POLL_SECONDS * 1000, self.poll_for_changes)

    def _build_ui(self):
        header = ttk.Frame(self, padding=12)
        header.pack(fill="x")
        ttk.Label(header, text="Trials Queue Dashboard", font=("Segoe UI", 18, "bold")).pack(side="left")
        ttk.Label(header, text=f"Excel: {self.store.excel_path}").pack(side="right")

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self.prep_tab = ttk.Frame(notebook, padding=12)
        self.coord_tab = ttk.Frame(notebook, padding=12)
        notebook.add(self.prep_tab, text="Pre-prep team")
        notebook.add(self.coord_tab, text="Device coordinators")
        self._build_preprep_tab()
        self._build_coordinator_tab()

    def _build_preprep_tab(self):
        form = ttk.LabelFrame(self.prep_tab, text="Add workbook line item", padding=12)
        form.pack(fill="x")
        self.prep_fields = {}
        labels = [
            ("first_name", "First name"),
            ("last_name", "Last name"),
            ("device", "Device"),
            ("loan_type", "Loan type"),
            ("queue_date", "Queue date"),
            ("vocabulary", "Vocabulary"),
        ]
        for index, (field, label) in enumerate(labels):
            ttk.Label(form, text=label).grid(row=index // 3 * 2, column=index % 3, sticky="w", padx=4)
            variable = StringVar()
            ttk.Entry(form, textvariable=variable, width=30).grid(row=index // 3 * 2 + 1, column=index % 3, sticky="ew", padx=4, pady=(0, 8))
            self.prep_fields[field] = variable
        ttk.Label(form, text="Priority").grid(row=4, column=0, sticky="w", padx=4)
        priority = StringVar(value="")
        ttk.Combobox(form, textvariable=priority, values=["", "EXPEDITE", "Funded Rental", "Accessories", "Requested Ship Date"], width=28).grid(row=5, column=0, sticky="ew", padx=4)
        self.prep_fields["priority"] = priority
        self.ready_for_prep = BooleanVar(value=True)
        ttk.Checkbutton(form, text="Ready for Prep (writes X)", variable=self.ready_for_prep).grid(row=5, column=1, sticky="w", padx=4)
        ttk.Button(form, text="Add to Excel", command=self.add_preprep_row).grid(row=5, column=2, sticky="e", padx=4)
        ttk.Label(form, text="Notes").grid(row=6, column=0, sticky="w", padx=4)
        self.notes = ScrolledText(form, width=88, height=4)
        self.notes.grid(row=7, column=0, columnspan=3, sticky="ew", padx=4)
        for column in range(3):
            form.columnconfigure(column, weight=1)

        self.prep_tree = self._make_tree(self.prep_tab, "Workbook preview")

    def _build_coordinator_tab(self):
        controls = ttk.LabelFrame(self.coord_tab, text="Queue controls", padding=12)
        controls.pack(fill="x")
        ttk.Label(controls, text="Your name / initials").pack(side="left")
        ttk.Entry(controls, textvariable=self.user_name, width=24).pack(side="left", padx=8)
        ttk.Button(controls, text="Join prep FIFO", command=lambda: self.join_queue("prep")).pack(side="left", padx=4)
        ttk.Button(controls, text="Join 2nd-check FIFO", command=lambda: self.join_queue("qa")).pack(side="left", padx=4)
        ttk.Button(controls, text="Refresh", command=self.refresh_all).pack(side="right")

        queue_frame = ttk.Frame(self.coord_tab)
        queue_frame.pack(fill="x", pady=10)
        self.prep_queue_label = ttk.Label(queue_frame, text="Prep FIFO: empty", font=("Segoe UI", 10, "bold"))
        self.prep_queue_label.pack(side="left", padx=(0, 20))
        self.qa_queue_label = ttk.Label(queue_frame, text="2nd-check FIFO: empty", font=("Segoe UI", 10, "bold"))
        self.qa_queue_label.pack(side="left")

        self.dashboard_tree = self._make_tree(self.coord_tab, "Ready coordinator work")
        actions = ttk.Frame(self.coord_tab)
        actions.pack(fill="x", pady=8)
        ttk.Button(actions, text="Offer prep to next", command=lambda: self.offer_selected("prep")).pack(side="left", padx=4)
        ttk.Button(actions, text="Claim prep", command=lambda: self.claim_selected("prep")).pack(side="left", padx=4)
        ttk.Button(actions, text="Prep complete", command=lambda: self.complete_selected("prep")).pack(side="left", padx=4)
        ttk.Button(actions, text="Skip prep", command=lambda: self.skip_selected("prep")).pack(side="left", padx=4)
        ttk.Separator(actions, orient="vertical").pack(side="left", fill="y", padx=12)
        ttk.Button(actions, text="Offer 2nd check to next", command=lambda: self.offer_selected("qa")).pack(side="left", padx=4)
        ttk.Button(actions, text="Claim 2nd check", command=lambda: self.claim_selected("qa")).pack(side="left", padx=4)
        ttk.Button(actions, text="2nd check complete", command=lambda: self.complete_selected("qa")).pack(side="left", padx=4)
        ttk.Button(actions, text="Skip 2nd check", command=lambda: self.skip_selected("qa")).pack(side="left", padx=4)

    def _make_tree(self, parent, title):
        frame = ttk.LabelFrame(parent, text=title, padding=8)
        frame.pack(fill="both", expand=True, pady=10)
        columns = ("lane", "id", "name", "device", "loan", "priority", "prep", "qa", "assigned", "status")
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=14)
        headings = {
            "lane": "Lane",
            "id": "ID",
            "name": "Name",
            "device": "Device",
            "loan": "Loan",
            "priority": "Priority",
            "prep": "Prep ready",
            "qa": "2nd check ready",
            "assigned": "Assigned",
            "status": "Status",
        }
        for column, heading in headings.items():
            tree.heading(column, text=heading)
            tree.column(column, width=120 if column != "device" else 180)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        return tree

    def add_preprep_row(self):
        try:
            payload = {field: variable.get().strip() for field, variable in self.prep_fields.items()}
            payload["notes"] = self.notes.get("1.0", "end").strip()
            payload["ready_for_prep"] = self.ready_for_prep.get()
            if not payload.get("first_name") or not payload.get("last_name") or not payload.get("device"):
                raise ValueError("First name, last name, and device are required.")
            row_id = self.store.add_row(payload)
            messagebox.showinfo("Saved", f"Added row {row_id} to Excel.")
            for variable in self.prep_fields.values():
                variable.set("")
            self.ready_for_prep.set(True)
            self.notes.delete("1.0", "end")
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("Could not add row", str(exc))

    def join_queue(self, queue_name):
        try:
            self.queues.join(queue_name, clean_user(self.user_name.get()))
            self.refresh_queues()
        except Exception as exc:
            messagebox.showerror("Queue error", str(exc))

    def selected_row_number(self):
        selection = self.dashboard_tree.selection()
        if not selection:
            raise ValueError("Select an order first.")
        return self.selected_rows[selection[0]]["row_number"]

    def offer_selected(self, queue_name):
        try:
            row_number = self.selected_row_number()
            row = self.store._find_dashboard_row(row_number)
            if not row:
                raise ValueError("Order is no longer available.")
            if row.get("assignment_active"):
                messagebox.showinfo("Offer active", f"Existing offer is still active for {row.get('assignment_user')}.")
                return
            previous_user = row.get("assignment_user")
            previous_status = str(row.get("assignment_status") or "")
            if previous_user and previous_status.startswith("Offered to"):
                self.queues.rotate_to_bottom(queue_name, previous_user)
                self.store.clear_assignment(row_number, previous_user, "Expired")
            assigned = self.queues.first_user(queue_name)
            if not assigned:
                raise ValueError("No coordinators are currently in this FIFO queue.")
            expires = datetime.now(timezone.utc) + timedelta(seconds=CLAIM_TIMEOUT_SECONDS)
            self.store.set_assignment(row_number, assigned, expires.replace(microsecond=0).isoformat(), f"Offered to {assigned}")
            messagebox.showinfo("Offer sent", f"Offer sent to {assigned} for {CLAIM_TIMEOUT_SECONDS} seconds.")
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("Offer error", str(exc))

    def claim_selected(self, queue_name):
        try:
            user = clean_user(self.user_name.get())
            self.store.claim_order(self.selected_row_number(), queue_name, user)
            messagebox.showinfo("Claimed", "Order claimed and Excel updated.")
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("Claim error", str(exc))

    def complete_selected(self, queue_name):
        try:
            user = clean_user(self.user_name.get())
            self.store.complete_order(self.selected_row_number(), queue_name, user)
            self.queues.rotate_to_bottom(queue_name, user)
            messagebox.showinfo("Completed", "Order completed and Excel updated.")
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("Complete error", str(exc))

    def skip_selected(self, queue_name):
        try:
            user = clean_user(self.user_name.get())
            self.queues.rotate_to_bottom(queue_name, user)
            self.store.clear_assignment(self.selected_row_number(), user, "Skipped")
            messagebox.showinfo("Skipped", "You were moved to the bottom of the FIFO queue.")
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("Skip error", str(exc))

    def refresh_all(self):
        try:
            rows = self.store.read_rows()
            self._fill_tree(self.prep_tree, rows)
            self._fill_tree(self.dashboard_tree, [row for row in rows if row["prep_ready"] or row["qa_ready"]])
            self.refresh_queues()
            self.notify_current_user()
        except Exception as exc:
            messagebox.showerror("Refresh error", str(exc))

    def refresh_queues(self):
        state = self.queues.load()
        self.prep_queue_label.config(text=f"Prep FIFO: {self._queue_text(state['prep'])}")
        self.qa_queue_label.config(text=f"2nd-check FIFO: {self._queue_text(state['qa'])}")

    def notify_current_user(self):
        user = self.user_name.get().strip()
        if not user:
            return
        for row in self.store.dashboard_rows():
            key = (row["row_number"], row.get("assignment_expires"))
            if row.get("assignment_user") == user and row.get("assignment_active") and key not in self.notified_offers:
                self.notified_offers.add(key)
                messagebox.showinfo("Order ready to claim", f"{row.get('display_name') or row.get('id')} is offered to you. Claim within {CLAIM_TIMEOUT_SECONDS} seconds.")

    def _fill_tree(self, tree, rows):
        tree.delete(*tree.get_children())
        if tree is self.dashboard_tree:
            self.selected_rows = {}
        for row in sorted(rows, key=sort_key):
            item_id = tree.insert(
                "",
                "end",
                values=(
                    row.get("lane"),
                    row.get("id") or row.get("row_number"),
                    row.get("display_name"),
                    row.get("device"),
                    row.get("loan_type"),
                    row.get("priority") or "Regular",
                    "Yes" if row.get("prep_ready") else "No",
                    "Yes" if row.get("qa_ready") else "No",
                    row.get("assignment_user") or "",
                    row.get("assignment_status") or "Ready",
                ),
            )
            if tree is self.dashboard_tree:
                self.selected_rows[item_id] = row

    @staticmethod
    def _queue_text(entries):
        return " → ".join(entry["user"] for entry in entries) if entries else "empty"

    def poll_for_changes(self):
        self.refresh_all()
        self.after(POLL_SECONDS * 1000, self.poll_for_changes)


def main():
    app = TrialsDashboard(ExcelQueueStore(EXCEL_PATH), QueueState(STATE_PATH))
    app.mainloop()


if __name__ == "__main__":
    main()
