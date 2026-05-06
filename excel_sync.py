"""Excel import/export helpers for the Daily Queue workbook.

Column mapping is intentionally kept at the top of this file so operations can
adjust to workbook changes without touching the rest of the app.  The app only
writes the mapped prep and QA cells for the selected row; it does not generate a
new workbook or overwrite unrelated cells.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from models import SECTION_DEFAULT

# ----------------------------- Excel mapping -----------------------------
# Sheet to import from.  If it is missing, the first worksheet is used.
QUEUE_SHEET_NAME = "Daily Queue"
HEADER_ROW = 1
FIRST_DATA_ROW = 2

# Source columns read from the Daily Queue workbook.
COLUMN_MAP = {
    "last_name": "A",
    "first_name": "B",
    "device": "C",
    "loan_type": "D",
    "queue_date": "E",
    "vocabulary": "F",
    "notes": "G",
}

# Destination columns updated by button actions per the MVP requirements.
PREP_USERNAME_COLUMN = "I"
QA_DONE_COLUMN = "J"
QA_DONE_SUFFIX = "-Yes"
# -------------------------------------------------------------------------


def _cell_text(value: Any) -> str:
    """Normalize Excel cell values for display and storage."""
    if value is None:
        return ""
    return str(value).strip()


def _is_section_row(first_cell: str) -> bool:
    """Detect dashed separator rows from the existing queue workbook."""
    return bool(first_cell) and "---" in first_cell


def _section_name(first_cell: str) -> str:
    """Convert a dashed Excel separator into a readable dashboard section."""
    return first_cell.replace("-", " ").strip() or SECTION_DEFAULT


def import_jobs_from_workbook(workbook_path: str | Path) -> list[dict]:
    """Read queue rows from Excel and return dictionaries ready for SQLite."""
    path = Path(workbook_path).expanduser().resolve()
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[QUEUE_SHEET_NAME] if QUEUE_SHEET_NAME in workbook.sheetnames else workbook.worksheets[0]

    jobs: list[dict] = []
    current_section = SECTION_DEFAULT
    for row_idx in range(FIRST_DATA_ROW, sheet.max_row + 1):
        first_cell = _cell_text(sheet[f"A{row_idx}"].value)
        if _is_section_row(first_cell):
            current_section = _section_name(first_cell)
            continue

        values = {field: _cell_text(sheet[f"{column}{row_idx}"].value) for field, column in COLUMN_MAP.items()}
        if not any(values.values()):
            continue

        jobs.append(
            {
                "workbook_path": str(path),
                "sheet_name": sheet.title,
                "excel_row": row_idx,
                "section": current_section,
                **values,
            }
        )
    workbook.close()
    return jobs


def write_prep_username(workbook_path: str | Path, sheet_name: str, excel_row: int, initials: str) -> None:
    """Write the prep username to the mapped Excel column for one row only."""
    _write_single_cell(workbook_path, sheet_name, excel_row, PREP_USERNAME_COLUMN, initials)


def write_qa_done(workbook_path: str | Path, sheet_name: str, excel_row: int, initials: str) -> None:
    """Write the QA completion marker to the mapped Excel column for one row only."""
    _write_single_cell(workbook_path, sheet_name, excel_row, QA_DONE_COLUMN, f"{initials}{QA_DONE_SUFFIX}")


def _write_single_cell(
    workbook_path: str | Path,
    sheet_name: str,
    excel_row: int,
    column: str,
    value: str,
) -> None:
    """Load the workbook, update a single mapped cell, and save in place."""
    path = Path(workbook_path).expanduser().resolve()
    workbook = load_workbook(path)
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[QUEUE_SHEET_NAME]
    sheet[f"{column}{excel_row}"] = value
    workbook.save(path)
    workbook.close()
